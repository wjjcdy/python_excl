#!/usr/bin/env python 
#coding:utf-8��
import openpyxl
import gc

DT = 200          # 时间窗口设置，单位ms，可依据需求修改
time_start = 1   # 时间起始参数

USID_MAX = 1000  # 一次最大处理量

Block_index = 1

# 读取需要处理的源数据excl
# 第一步：打开工作簿
print("打开源文件等待中......")
source_data = openpyxl.load_workbook('source_test.xlsx')
# 第二步：选取表单
sh = source_data['source_sheet']
# 第三步：获取数据个数
size_line = len(list(sh.rows)) - 1
#调试打印多少行
print("打开文件，共有"+ str(size_line) + "行数据")
#初始化当前处理的是第几个周期
time_block = 1

#初始化数据结果统计表
dict_userID = {}

#初始化userID对应excl列数索引表
cols_userID = {}
userID_set = set()       # 集合
userID_list = []         # 有序列表
userID_list_block = []   # 列表块

#初始化写入行数
row_index_write = 2

time_start = list(sh.rows)[1][0].value
#数据批量处理
############################################################
# 数据预处理，将整个表分块处理
############################################################
print("数据分析中，请稍等......" )
for cases in list(sh.rows)[1:]: 
    #获取每一行所有元素
    time_stamp =  cases[0].value - time_start +1
    LcID = cases[1].value
    Ueid = cases[2].value
    bufferSize = cases[3].value
    
    userID_set.add(Ueid)    
block_num = len(userID_set) // USID_MAX +1                     # 计算几个块处理
userID_list = list(userID_set)                                  # 转换成有序序列
userID_set.clear()
del userID_set
gc.collect()

for i in range(block_num-1):
    temp_list = userID_list[i*USID_MAX:(i+1)*USID_MAX]           # 切片每块
    for j,id in enumerate(temp_list):
        cols_userID[id] = j +2  
    userID_list_block.append(set(temp_list))                     # 每块成一个集合，放入集合块

if len(userID_list) % USID_MAX:
    temp_list = userID_list[(block_num-1)*USID_MAX:]             # 若不能整除，表明留有最后一块
    for j,id in enumerate(temp_list):
        cols_userID[id] = j +2  
    userID_list_block.append(set(temp_list))                     # 每块成一个集合，放入集合块
 
print("分析完成,Useid个数：" + str(len(userID_list)))
del userID_list
gc.collect()
print("分成" + str(block_num)+"块进行处理: "+str(len(userID_list_block)))
############################################################
# 将每块内所有useid进行处理
############################################################
for i in range(len(userID_list_block)):
    # 创建一个result_sheet的sheet表单
    sheet_name = 'result_sheet' + str(i)
    # 创建一个工作簿，用于存储结果
    result_data = openpyxl.Workbook()
    # 创建一个result_sheet的sheet表单
    result_data.create_sheet('result_sheet')
    result_sheet = result_data['result_sheet']
    # 填写表头注释，即第1行，第1列，写入
    result_sheet.cell(1, 1).value = "period("+ str(DT)+"ms)/userID"
    print("第"+str(i)+"块数据开始处理，请稍等...")
    #数据批量处理
    for cases in list(sh.rows)[1:]: 
        #获取每一行所有元素
        time_stamp =  cases[0].value - time_start +1
        LcID = cases[1].value
        Ueid = cases[2].value
        bufferSize = cases[3].value

        if userID_list_block[i].__contains__(Ueid):        # Ueid 在此块中则处理
        #################################################################################
        #一个周期结束统计处理
        #################################################################################
            if time_stamp > time_block*DT:                                             # 当每一周期时间结束进行统计处理
                result_sheet.cell(row_index_write, 1).value = "Period" + str(row_index_write-1)
                for useid_inf in dict_userID:                                          # 遍历此周期内所有出现的userID
                                                                                    # 第一列写入当前第几个周期
                    for lcid_inf in dict_userID[useid_inf]["data"]:                    # 遍历此周期内，此userID下所有lcID，并将其buffsize累加             
                        dict_userID[useid_inf]["sum"] += dict_userID[useid_inf]["data"][lcid_inf]["sum"]
                        dict_userID[useid_inf]["data"][lcid_inf]["sum"] = 0            # 初始化新的时间窗口此逻辑ID的buffer_sum为0

                    # 此周期中某一userID的size总和写入对应位置中
                    result_sheet.cell(row_index_write, cols_userID[useid_inf]).value = dict_userID[useid_inf]["sum"]
                    dict_userID[useid_inf]["sum"] = 0                                  #初始化新的时间窗口，此useid的sum清零初始化
                row_index_write +=1
                #time_block = time_block + 1      # 当前周期处理完成，更新周期序号
                time_block = time_stamp//DT + 1
            
        #################################################################################3
        # 周期内进行一行行数据处理    
        #################################################################################
        
            #遇到新的Userid，创建一个userID结果列表
            if not (dict_userID.__contains__(Ueid)):
                dict_userID.setdefault(Ueid, {"data":{},"sum":0})
                result_sheet.cell(1, cols_userID[Ueid]).value = Ueid     
                    
            #当前Userid列表内遇到新的LcID，创建LcID子列表
            if not (dict_userID[Ueid]["data"].__contains__(LcID)):
                dict_userID[Ueid]["data"].setdefault(LcID, {"last":0,"sum":0})
                dict_userID[Ueid]["data"][LcID]["last"] = bufferSize                 #初始化useid 下 locid的buffsize，不做统计
                dict_userID[Ueid]["data"][LcID]["sum"] = 0                           #初始化本locid 的sum
            else:                                                                    
                size_increase = bufferSize - dict_userID[Ueid]["data"][LcID]["last"] #若已存在，则应计算增加量
                dict_userID[Ueid]["data"][LcID]["sum"] += size_increase              #统计一个窗口的累加和

            # 记录上时刻，某一Ueid，某一LcID下的buffeSize
            dict_userID[Ueid]["data"][LcID]["last"] = bufferSize

    #################################################################################
    #最后一个周期结束统计处理，原因是最后一次不一定是整周期，故以上处理会遗漏最后一周期数据
    #################################################################################   
    result_sheet.cell(row_index_write, 1).value = "Period" + str(row_index_write-1) 
    for useid_inf in dict_userID:
        for lcid_inf in dict_userID[useid_inf]["data"]:
            dict_userID[useid_inf]["sum"] += dict_userID[useid_inf]["data"][lcid_inf]["sum"]
        # 写入结果的excl表格中
        result_sheet.cell(row_index_write, cols_userID[useid_inf]).value = dict_userID[useid_inf]["sum"]

    #初始化数据库和数据
    dict_userID.clear()
    
    row_index_write = 2
    time_block = 1
    print("第"+str(i)+"块数据处理完成，文件保存中......")
    result_data.save(sheet_name+'.xlsx')
    result_data.close()
    print(sheet_name+'.xlsx文件生成')

    del result_data
    del result_sheet
    gc.collect()

print("全部数据处理完成")
# 关闭工作薄
source_data.close()

# # 将新创建的结果工作蒲保存为一个xlsx格式的文件
# result_data.save('result.xlsx')


